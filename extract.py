from openpyxl import Workbook,load_workbook
import requests
from math import log10
import csv
import textwrap

def getWinProbability(ra,rb):
    return 1.0 / (1.0 + pow( 10.0, (rb - ra) / 400.0))

def aggregateRatings(teamRatings):
    if(len(teamRatings)==0):
        return 0

    left = 1
    right = 1E4

    for tt in range(100):
        r = (left + right) / 2.0

        rWinsProbability = 1.0
        for i in range (len(teamRatings)):
            rWinsProbability *= getWinProbability(r, teamRatings[i])

        rating = log10(1 / (rWinsProbability) - 1) * 400 + r

        if (rating > r):
            left = r
        else:
            right = r

    return (left + right) / 2.0

color=[
{'title':'Legendary grandmaster','user':'user-legendary','lo':3000,'hi':3999},
{'title':'International Grandmaster','user':'user-red','lo':2600,'hi':2999},
{'title':'Grandmaster','user':'user-red','lo':2400,'hi':2599},
{'title':'International master','user':'user-orange','lo':2300,'hi':2399},
{'title':'Master','user':'user-orange','lo':2100,'hi':2299},
{'title':'Candidate master','user':'user-violet','lo':1900,'hi':2099},
{'title':'Expert','user':'user-blue','lo':1600,'hi':1899},
{'title':'Specialist','user':'user-cyan','lo':1400,'hi':1599},
{'title':'Pupil','user':'user-green','lo':1200,'hi':1399},
{'title':'Newbie','user':'user-gray','lo':0,'hi':1199},
]

def getColor(rating):
    for i in range(len(color)):
        if(rating>=color[i]['lo'] and rating<=color[i]['hi']):
            return color[i]

def reduce(text):
    ntxt=textwrap.wrap(text, 11, break_long_words=True)
    for i in range(len(ntxt)):
        if ntxt[i][0]=='_':
            ntxt[i]='\\'+ntxt[i]
            # print(ntxt[i])
    return " ".join(ntxt)

def getLink(user):
    col=getColor(user['rate'])
    link='<a class="rated-user '+col['user']+'" href="/profile/'+user['han']+'" title="'+col['title']+' '+ user['han']+'">'+reduce(user['han'])+'</a>'
    return link

def getRating(handle):
    URL = "https://codeforces.com/api/user.info"
    PARAMS = {'handles': handle}
    while True:
        try:
            res = requests.get(url=URL, params=PARAMS,timeout=10).json()
            if (res['status'] == 'OK'):
                if 'maxRating' in res['result'][0]:
                    return res['result'][0]['maxRating']
                else:
                    return 0
            else:
                return 0
        except Exception as e:
            print(e)
    return 0

wb = load_workbook("icpcgp.xlsx")
ws = wb.active
teamList=[]
for i in range (2,129):
    print("pre "+str(i))
    mem=[{'han':ws.cell(i,4).value,'rate':0},{'han':ws.cell(i,5).value,'rate':0},{'han':ws.cell(i,6).value,'rate':0}]
    rat=0
    rated=[]
    for j in range(3):
        if mem[j]['han']:
            print(mem[j]['han'], end=" ")
            mem[j]['rate']=getRating(mem[j]['han'])
            mem[j]['han']=mem[j]['han']
            rated.append(mem[j]['rate'])
    mem = sorted(mem, key=lambda k: k['rate'],reverse=True)
    rat=round(aggregateRatings(rated))
    print(rat)
    team={
        'air':int(ws.cell(i,1).value),
        'name':ws.cell(i,2).value,
        'inst':ws.cell(i,3).value,
        'mem1':mem[0],
        'mem2':mem[1],
        'mem3':mem[2],
        'loc':ws.cell(i,7).value,
        'rat':rat
    }
    # print(team)
    teamList.append(team)

teamList = sorted(teamList, key=lambda k: (k['rat'],-k['air']),reverse=True)
with open('result.csv', 'w') as csvFile:
    writer = csv.writer(csvFile)
    header=['No.','AIR','Team Name','Institute Name','Member 1','Member 2','Member 3','Rating','Location']
    writer.writerow(header)
    for i in range(len(teamList)):
        # print(teamList[i])
        if teamList[i]['mem1']['rate']>0:
            teamList[i]['mem1']['han']=getLink(teamList[i]['mem1'])
        if teamList[i]['mem2']['rate']>0:
            teamList[i]['mem2']['han']=getLink(teamList[i]['mem2'])
        if teamList[i]['mem3']['rate']>0:
            teamList[i]['mem3']['han']=getLink(teamList[i]['mem3'])
        if teamList[i]['rat']:
            teamList[i]['name']=getLink({'han':teamList[i]['name'],'rate':teamList[i]['rat']})
            teamList[i]['rat']=getLink({'han':str(teamList[i]['rat']),'rate':teamList[i]['rat']})
        row=[i+1]
        row.append(teamList[i]['air'])
        row.append(teamList[i]['name'])
        row.append(teamList[i]['inst'])
        row.append(teamList[i]['mem1']['han'])
        row.append(teamList[i]['mem2']['han'])
        row.append(teamList[i]['mem3']['han'])
        row.append(teamList[i]['rat'])
        row.append(teamList[i]['loc'])
        writer.writerow(row)